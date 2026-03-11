"""
Rutas de productos: búsqueda por código de barras, importación CSV, exportación Excel.
"""

import csv
import io
import logging
from typing import Optional

from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel
from typing import List

from models.schemas import Producto, ProductoRespuesta

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/productos", tags=["Productos"])


# ------------------------------------------------------------------ #
#  Referencia al cliente Odoo (se inyecta desde main.py)               #
# ------------------------------------------------------------------ #
_odoo_client = None


def set_odoo_client(client) -> None:
    """Inyecta la instancia del cliente Odoo."""
    global _odoo_client
    _odoo_client = client


# ------------------------------------------------------------------ #
#  Rutas                                                               #
# ------------------------------------------------------------------ #

@router.get("/buscar", response_model=ProductoRespuesta)
async def buscar_producto(codigo_barras: str):
    """Busca un producto en Odoo por su código de barras."""
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")

    try:
        resultado = _odoo_client.buscar_producto_por_codigo(codigo_barras)
        if resultado:
            producto = Producto(
                id_odoo=resultado["id"],
                codigo_barras=resultado.get("barcode", codigo_barras),
                nombre=resultado.get("name", "Sin nombre"),
                cantidad=0,
                precio=resultado.get("precio", 0.0),
            )
            return ProductoRespuesta(
                encontrado=True,
                producto=producto,
                mensaje=f"Producto encontrado: {producto.nombre}",
            )
        return ProductoRespuesta(
            encontrado=False,
            mensaje=f"Producto no encontrado para código: {codigo_barras}",
        )
    except Exception as e:
        logger.error("Error buscando producto: %s", e)
        raise HTTPException(status_code=500, detail=f"Error al buscar producto: {str(e)}")

class BusquedaMasivaRequest(BaseModel):
    codigos: List[str]

@router.post("/buscar-masivo")
async def buscar_masivo(req: BusquedaMasivaRequest):
    """Busca múltiples productos en Odoo por sus códigos de barras de una sola vez."""
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")

    try:
        productos = _odoo_client.buscar_productos_por_codigos(req.codigos)
        return {"exito": True, "productos": productos}
    except Exception as e:
        logger.error("Error en búsqueda masiva: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/stock")
async def obtener_stock(id_odoo: int):
    """
    Obtiene la cantidad a la mano en todas las ubicaciones internas
    y la fecha del último movimiento de un producto.
    """
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")

    try:
        stock = _odoo_client.obtener_stock_a_la_mano(id_odoo)
        ultimo_movimiento = _odoo_client.obtener_ultimo_movimiento(id_odoo)
        return {
            "id_odoo": id_odoo,
            "stock_por_ubicacion": stock,
            "ultimo_movimiento": ultimo_movimiento,
        }
    except Exception as e:
        logger.error("Error obteniendo stock: %s", e)
        raise HTTPException(status_code=500, detail=f"Error al obtener stock: {str(e)}")


@router.post("/stock-masivo")
async def obtener_stock_masivo(request: dict):
    """
    Obtiene la cantidad a la mano masiva para una lista de product_ids.
    Cuerpo esperado: {"ids_odoo": [123, 456]}
    """
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")

    ids_odoo = request.get("ids_odoo", [])
    if not ids_odoo:
        return {"resultados": {}}

    try:
        stock_map = _odoo_client.obtener_stock_masivo(ids_odoo)
        return {"resultados": stock_map}
    except Exception as e:
        logger.error("Error obteniendo stock masivo: %s", e)
        raise HTTPException(status_code=500, detail=f"Error al obtener stock masivo: {str(e)}")


@router.get("/categorias")
async def obtener_categorias():
    """Devuelve todas las categorías de productos disponibles en Odoo."""
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")
    try:
        return {"categorias": _odoo_client.obtener_categorias()}
    except Exception as e:
        logger.error("Error obteniendo categorías: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/por-categoria")
async def obtener_por_categoria(categoria_id: int):
    """
    Devuelve los productos de una categoría con su stock actual
    en todas las ubicaciones internas.
    """
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")
    try:
        productos = _odoo_client.obtener_productos_por_categoria(categoria_id)
        return {"categoria_id": categoria_id, "productos": productos, "total": len(productos)}
    except Exception as e:
        logger.error("Error obteniendo productos por categoría: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buscar-global")
async def buscar_global(q: str):
    """
    Busca productos en todo el almacén por nombre o código y devuelve su stock.
    """
    if not _odoo_client:
        raise HTTPException(status_code=503, detail="Cliente Odoo no configurado")
    try:
        productos = _odoo_client.buscar_productos_global(q)
        return {"query": q, "productos": productos, "total": len(productos)}
    except Exception as e:
        logger.error("Error en búsqueda global: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/importar-csv")
async def importar_csv(archivo: UploadFile = File(...)):
    """
    Importa productos desde un archivo CSV.
    Formato esperado: codigo_barras,nombre (con encabezados).
    """
    if not archivo.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un CSV")

    contenido = await archivo.read()
    texto = contenido.decode("utf-8-sig")
    lector = csv.DictReader(io.StringIO(texto))

    productos: list[dict] = []
    for fila in lector:
        codigo = fila.get("codigo_barras") or fila.get("barcode") or ""
        nombre = fila.get("nombre") or fila.get("name") or ""
        if codigo.strip():
            productos.append({"codigo_barras": codigo.strip(), "nombre": nombre.strip()})

    logger.info("CSV importado: %d productos", len(productos))
    return {"mensaje": f"Se importaron {len(productos)} productos", "productos": productos}


@router.post("/exportar-excel")
async def exportar_excel(items: list[Producto]):
    """Exporta la lista de productos contados a un archivo Excel (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo de Inventario"

    # Encabezados
    encabezados = ["Código de Barras", "Producto", "Cantidad"]
    ws.append(encabezados)

    # Estilo de encabezados
    from openpyxl.styles import Font, PatternFill
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    # Datos
    for item in items:
        ws.append([item.codigo_barras, item.nombre, item.cantidad])

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=conteo_inventario.xlsx"},
    )
